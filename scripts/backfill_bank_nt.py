#!/usr/bin/env python3
"""Novčani tok banaka iz NT_D lista nadzornog godišnjeg obrasca (1Y XLSX).

Banke dosad nisu imale operating/investing/financing CF (parser je čitao samo
RDG + bilancu) pa je grupa "Novčani tok" na Pokazateljima bila n/p. NT_D list
obrasca ima direktnu+indirektnu metodu s AOP oznakama:
  AOP 34 = neto NT iz poslovnih, 40 = ulagačkih, 47 = financijskih,
  45 = (Isplaćena dividenda), 35 = primici/plaćanja za materijalnu imovinu.

Dual-match (AOP + naziv retka) i sanity gate AOP48 ≈ 34+40+47 — datoteka koja
ne prođe se preskače s razlogom (radije rupa nego kriva brojka). Upis na
POSTOJEĆI annual/consolidated filing (DELETE+INSERT po itemu, idempotentno).

Pokretanje:  python -m scripts.backfill_bank_nt
"""
from __future__ import annotations

import os
import re
import sys
import tempfile

sys.path.insert(0, ".")

from openpyxl import load_workbook  # noqa: E402

from src.db import get_conn  # noqa: E402

LOCAL_DIR = "/tmp/bank_xlsx"

# (item, AOP, obavezni dio naziva retka, uzmi_abs, samo_ako_negativno)
NT_MAP = [
    ("operating_cf", "34", r"poslovnih aktivnosti", False, False),
    ("investing_cf", "40", r"ulaga[čc]kih aktivnosti", False, False),
    ("financing_cf", "47", r"financijskih aktivnosti", False, False),
    ("dividends_paid", "45", r"ispla[ćc]ena dividenda", True, False),
    ("capex", "35", r"kupnju ma", True, True),  # samo neto ODLJEV = capex
]


def _rows(ws) -> dict[str, tuple[str, float | None]]:
    """{AOP: (label, tekuća_vrijednost)} — stupci: 0 naziv, 6 AOP, 8 tekuće."""
    out = {}
    for row in ws.iter_rows(min_row=1, max_row=120, max_col=10, values_only=True):
        label, aop = row[0], row[6] if len(row) > 6 else None
        cur = row[8] if len(row) > 8 else None
        if label is None or aop is None:
            continue
        try:
            val = float(cur) if cur is not None else None
        except (TypeError, ValueError):
            val = None
        out[str(aop).strip()] = (str(label).strip(), val)
    return out


def parse_nt(path: str) -> dict[str, float] | None:
    wb = load_workbook(path, read_only=True, data_only=True)
    name = next((s for s in wb.sheetnames if s.upper().startswith("NT")), None)
    if not name:
        return None
    rows = _rows(wb[name])
    got: dict[str, float] = {}
    for item, aop, label_rx, take_abs, only_neg in NT_MAP:
        r = rows.get(aop)
        if not r or not re.search(label_rx, r[0], re.I) or r[1] is None:
            continue  # dual-match pao ili prazno -> preskoči item
        v = r[1]
        if only_neg and v >= 0:
            continue  # neto PRODAVATELJ imovine -> capex nije definiran
        got[item] = abs(v) if take_abs else v
    # sanity: neto promjena novca (48) = 34+40+47 (tolerancija 1% ili 1000 EUR)
    tot = rows.get("48")
    parts = [got.get("operating_cf"), got.get("investing_cf"), got.get("financing_cf")]
    if tot and tot[1] is not None and None not in parts:
        s = sum(parts)
        if abs(s - tot[1]) > max(abs(tot[1]) * 0.01, 1000):
            print(f"  sanity pao: Σ(34,40,47)={s:,.0f} vs AOP48={tot[1]:,.0f}")
            return None
    return got or None


def main() -> int:
    ok = skipped = 0
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""SELECT c.id, c.ticker, f.id, f.fiscal_year, f.source_url
                       FROM companies c JOIN filings f ON f.company_id=c.id
                       WHERE c.sector IN ('bank','insurance')
                         AND f.period_type='annual' AND f.basis='consolidated'
                         AND f.fiscal_year = (SELECT MAX(f2.fiscal_year) FROM filings f2
                              WHERE f2.company_id=c.id AND f2.period_type='annual'
                                AND f2.basis='consolidated')
                         AND NOT EXISTS (SELECT 1 FROM financials fin
                              WHERE fin.filing_id=f.id AND fin.item='operating_cf')
                       ORDER BY c.ticker""")
        for cid, tick, fid, fy, url in cur.fetchall():
            path = os.path.join(LOCAL_DIR, f"{tick}.xlsx")
            if not os.path.exists(path):
                # probaj skinuti 1Y XLSX sa zabilježenog izvora (EHO)
                if url and url.lower().endswith(".xlsx"):
                    try:
                        import requests
                        r = requests.get(url, timeout=60)
                        r.raise_for_status()
                        path = os.path.join(tempfile.gettempdir(), f"nt_{tick}.xlsx")
                        with open(path, "wb") as fh:
                            fh.write(r.content)
                    except Exception as e:  # noqa: BLE001
                        print(f"[NT] {tick} FY{fy}: download pao ({str(e)[:60]}) -> preskačem")
                        skipped += 1
                        continue
                else:
                    print(f"[NT] {tick} FY{fy}: nema lokalnog XLSX ni .xlsx izvora -> preskačem")
                    skipped += 1
                    continue
            try:
                got = parse_nt(path)
            except Exception as e:  # noqa: BLE001
                print(f"[NT] {tick} FY{fy}: greška čitanja ({str(e)[:60]}) -> preskačem")
                skipped += 1
                continue
            if not got:
                print(f"[NT] {tick} FY{fy}: NT list nije nađen / gate pao -> preskačem")
                skipped += 1
                continue
            for item, val in got.items():
                cur.execute("DELETE FROM financials WHERE filing_id=%s AND item=%s",
                            (fid, item))
                cur.execute(
                    """INSERT INTO financials (filing_id, company_id, fiscal_year,
                         period_type, basis, statement, item, value_eur, confidence,
                         source_page, is_reported)
                       VALUES (%s,%s,%s,'annual','consolidated','cash_flow',%s,%s,
                               0.95,'NT list nadzornog obrasca (1Y XLSX)',TRUE)""",
                    (fid, cid, fy, item, val))
            conn.commit()
            ok += 1
            print(f"[NT] {tick} FY{fy}: " + ", ".join(
                f"{k}={v:,.0f}" for k, v in got.items()))
    print(f"\nGOTOVO: ok={ok}, preskočeno={skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
