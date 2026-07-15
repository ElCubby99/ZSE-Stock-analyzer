#!/usr/bin/env python3
"""Z2: povijest dividendi iz NT obrazaca (deterministički backfill).

Izvor: redak "Novčani izdaci za isplatu dividendi" (TFI NT_I AOP 41 / NT_D 35)
i "(Isplaćena dividenda)" (nadzorni bankovni NT_D AOP 45) — svaki obrazac
nosi TEKUĆU i PROŠLU godinu, pa 4Q/1Y datoteke više godina daju seriju.

Konvencije (dokumentirane uz svaki zapis):
- fiskalna godina dividende = godina isplate − 1 (GS u proljeće izglasa
  dividendu iz dobiti prethodne godine — standard na ZSE)
- iznos/dionici = ukupno isplaćeno / broj dionica DANAS (v_shares_canonical);
  aproksimacija označena u izvoru. SAMO za firme s JEDNOM klasom — kod više
  klasa raspodjela po klasi nije poznata iz NT-a pa se zapis preskače.
- ex/record/payment datumi NISU poznati iz NT-a -> NULL (prikaz "—").
- postojeći STVARNI zapis (ZSE/EHO) za istu fiskalnu godinu ima prednost —
  izvedeni se ne upisuje.

Pokretanje:  python -m scripts.backfill_dividends_nt
"""
from __future__ import annotations

import glob
import os
import re
import sys

sys.path.insert(0, ".")

from openpyxl import load_workbook  # noqa: E402

from src.db import get_conn  # noqa: E402

TFI_DIR = "/tmp/tfi_interim"
BANK_DIR = "/tmp/bank_xlsx"

DIV_RX = re.compile(r"izdaci za isplatu dividendi|ispla[ćc]ena dividenda", re.I)


def paid_from_xlsx(path: str) -> tuple[float | None, float | None] | None:
    """(isplaćeno_prošle, isplaćeno_tekuće) iz NT lista; None = nema retka."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return None
    for sn in wb.sheetnames:
        if not sn.upper().startswith("NT"):
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=120, max_col=12, values_only=True):
            label = next((str(v) for v in row if isinstance(v, str) and v.strip()), "")
            if not DIV_RX.search(label):
                continue
            nums = [v for v in row if isinstance(v, (int, float))]
            # [AOP, prošla, tekuća] — AOP je mali int; uzmi zadnja 2 broja
            vals = [v for v in nums if abs(v) > 200 or v == 0][-2:]
            if len(vals) == 2:
                prev_p, cur_p = abs(vals[0]), abs(vals[1])
                if prev_p or cur_p:
                    return (prev_p or None, cur_p or None)
    return None


def main() -> int:
    files = []  # (ticker, godina_izvještaja, path)
    for p in glob.glob(f"{TFI_DIR}/*_4Q.xlsx"):
        m = re.match(r"([A-Z0-9]+)_(\d{4})_4Q", os.path.basename(p))
        if m:
            files.append((m.group(1), int(m.group(2)), p))
    for p in glob.glob(f"{BANK_DIR}/*.xlsx"):
        files.append((os.path.basename(p).split(".")[0], 2025, p))

    ok = skipped = 0
    with get_conn() as conn, conn.cursor() as cur:
        for tick, ry, path in sorted(files):
            cur.execute("""SELECT c.id, COUNT(sc.id),
                                  MIN(sc.id) FILTER (WHERE sc.is_primary_line),
                                  MIN(sc.ticker) FILTER (WHERE sc.is_primary_line)
                           FROM companies c JOIN share_classes sc ON sc.company_id=c.id
                           WHERE c.ticker=%s GROUP BY c.id""", (tick,))
            r = cur.fetchone()
            if not r:
                continue
            cid, n_classes, scid, sctick = r
            if n_classes != 1:
                skipped += 1
                continue  # raspodjela po klasama nepoznata -> pošten preskok
            cur.execute("SELECT shares_ex_treasury FROM v_shares_canonical WHERE company_id=%s", (cid,))
            sr = cur.fetchone()
            shares = float(sr[0]) if sr and sr[0] else None
            if not shares:
                skipped += 1
                continue
            got = paid_from_xlsx(path)
            if not got:
                skipped += 1
                continue
            for pay_year, paid in ((ry - 1, got[0]), (ry, got[1])):
                if not paid or paid < 1000:  # simbolične/nula isplate preskačemo
                    continue
                fy = pay_year - 1  # konvencija: isplata u X = dividenda iz dobiti X-1
                dps = round(paid / shares, 4)
                # stvarni zapis za istu FY ima prednost (fiskalna godina
                # stvarnog zapisa: upisana ili ex_date.godina - 1)
                cur.execute("""SELECT 1 FROM dividends WHERE company_id=%s AND
                     COALESCE(fiscal_year, EXTRACT(YEAR FROM COALESCE(ex_date, payment_date))::int - 1) = %s
                     LIMIT 1""", (cid, fy))
                if cur.fetchone():
                    continue
                cur.execute(
                    """INSERT INTO dividends (company_id, share_class_id, class_ticker,
                         fiscal_year, amount_eur, div_type, source_url)
                       VALUES (%s,%s,%s,%s,%s,'izvedeno (NT obrazac)',%s)""",
                    (cid, scid, sctick, fy, dps,
                     f"{os.path.basename(path)}: NT redak 'isplaćene dividende' "
                     f"{paid:,.0f} EUR u {pay_year}. / {shares:,.0f} dionica danas "
                     f"(aproksimacija); fiskalna godina = godina isplate - 1"))
                ok += 1
                print(f"[div-NT] {tick}: FY{fy} ≈ {dps} €/dionici "
                      f"(isplaćeno {paid:,.0f} u {pay_year}.)")
        conn.commit()
    print(f"\nGOTOVO: izvedenih zapisa={ok}, preskočeno datoteka={skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
