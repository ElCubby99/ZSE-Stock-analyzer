"""M20: interim (kvartalni) backfill za BANKE i OSIGURANJE — 0 kredita.

Financijske firme objavljuju interim XLSX u TRI deterministička layouta:
  1. NADZORNI (isti kao godišnji malih banaka) — ZABA/HPB interim:
     parse_bank_tfi iz scripts/parse_bank_universe (AOP + label match).
  2. FINREP — kvartali malih banaka (IKBA/KBZ/PDBA/SNBA): drugi AOP-ovi;
     TOI je direktno AOP 017 ("Ukupno prihodi iz poslovanja, neto").
  3. OSIGURATELJSKI (IFRS 17; ISD/IFP listovi) — CROS: pozicije s tekstualnim
     opisom; prihod = "Prihodi od ugovora o osiguranju" (baza POTVRĐENA ista
     kao godišnja: Q1 152M x 4 ≈ FY 606,8M).

Svaki parser ima sanity gate (AOP + labela) — kriva forma vraća None i
preskače se S RAZLOGOM (ništa izmišljeno). Kolone: RDG kumulativ tekuće =
3. podatkovna od 4 (prior kum, prior Q, TEKUĆI KUM, tekući Q); bilanca =
zadnja od 2. Upis kroz validate gate kao q1/h1/9m/q4, cumulative=TRUE.

Pokretanje:  python -m scripts.backfill_fin_interim [--tickers ...]
"""
from __future__ import annotations

import argparse
import datetime
import os
import re
import sys

sys.path.insert(0, ".")

import openpyxl  # noqa: E402
import requests  # noqa: E402

from scripts.parse_bank_universe import _aop_rows, _get, parse_bank_tfi  # noqa: E402
from scripts.parse_tfi_universe import _verify  # noqa: E402
from src.db import get_conn  # noqa: E402
from src.loader import load_extraction  # noqa: E402
from src.eho import feed  # noqa: E402
from src.validator import validate_filing  # noqa: E402

PERIOD_MAP = {"1Q": "q1", "2Q": "h1", "1H": "h1", "3Q": "9m", "4Q": "q4"}
SCRATCH = "/tmp/fin_interim"


def parse_finrep(path: str) -> dict | None:
    """FINREP kvartalni layout (male banke): AOP 017 = TOI direktno."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "RDG" not in wb.sheetnames or "Bilanca" not in wb.sheetnames:
        return None
    rdg = _aop_rows(wb["RDG"])
    bil = _aop_rows(wb["Bilanca"])
    # sanity: FINREP RDG ima 'Kamatni prihodi' na AOP 1 i TOI na AOP 17
    if (_get(rdg, 1, r"^\s*Kamatni prihodi") is None
            or _get(rdg, 17, r"Ukupno prihodi iz poslovanja") is None
            or _get(bil, 32, r"Ukupna imovina") is None):
        return None
    items: dict[str, float] = {}
    src: dict[str, str] = {}

    def put(k, v, s):
        if v is not None:
            items[k] = v
            src[k] = s

    ii = _get(rdg, 1, r"Kamatni prihodi")
    ie = _get(rdg, 2, r"Kamatni rashodi")
    if ii is not None and ie is not None:
        put("net_interest_income", ii - ie, "FINREP RDG AOP 001−002")
    fi = _get(rdg, 5, r"naknada i provizija")
    fe = _get(rdg, 6, r"naknada i provizija")
    if fi is not None and fe is not None:
        put("net_fee_income", fi - fe, "FINREP RDG AOP 005−006")
    put("total_operating_income", _get(rdg, 17, r"Ukupno prihodi iz poslovanja"),
        "FINREP RDG AOP 017 (ukupno prihodi iz poslovanja, neto)")
    prov = [(_get(rdg, a, r"ezerviranja|manjenje vrijednosti") or 0.0)
            for a in (22, 23, 24, 25)]
    put("loan_loss_provisions", sum(prov),
        "FINREP RDG AOP 022+023+024+025 (rezerviranja + umanjenja)")
    pt = _get(rdg, 29, r"prije oporezivanja")
    pt2 = _get(rdg, 33, r"prije oporezivanja") or 0.0
    if pt is not None:
        put("pretax_income", pt + pt2, "FINREP RDG AOP 029(+033)")
    tx = _get(rdg, 30, r"[Pp]orezni")
    tx2 = _get(rdg, 34, r"[Pp]orezni") or 0.0
    if tx is not None:
        put("income_tax", tx + tx2, "FINREP RDG AOP 030(+034)")
    put("net_income", _get(rdg, 35, r"tekuće godine"), "FINREP RDG AOP 035")
    mi = _get(rdg, 36, r"manjinsk")
    if mi is not None:
        put("net_income_minority", mi, "FINREP RDG AOP 036")
    put("net_income_parent", _get(rdg, 37, r"matičnog društva"), "FINREP RDG AOP 037")
    put("total_assets", _get(bil, 32, r"Ukupna imovina"), "FINREP Bilanca AOP 032")
    beq = _get(bil, 67, r"Ukupno kapital")
    bmi = _get(bil, 66, r"Manjinski")
    put("total_equity", beq, "FINREP Bilanca AOP 067 (ukupno kapital)")
    if bmi is not None:
        put("minority_interests", bmi, "FINREP Bilanca AOP 066")
        if beq is not None:
            put("equity_parent", beq - bmi, "FINREP Bilanca AOP 067−066")
    return {"items": items, "src": src}


def _isd_rows(ws) -> dict[str, tuple[str, float | None]]:
    """Osigurateljski list: {broj_pozicije: (opis, tekući kumulativ 'Ukupno')}.
    Kolone: [broj, elementi, oznaka, opis, prior(Ž,N,UK), tekuće(Ž,N,UK)]."""
    out = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=16):
        vals = [c.value for c in row]
        pos = str(vals[0] or "").strip()
        if not re.fullmatch(r"\d{3}", pos):
            continue
        opis = str(vals[3] or "").strip()
        raw = vals[9] if len(vals) > 9 else None
        try:
            v = float(str(raw).replace(".", "").replace(",", ".")) \
                if isinstance(raw, str) and raw.strip() not in ("", "-") \
                else (float(raw) if isinstance(raw, (int, float)) else None)
        except ValueError:
            v = None
        if pos not in out:
            out[pos] = (opis, v)
    return out


def _isd_get(rows: dict, pos: str, label_rx: str) -> float | None:
    if pos not in rows:
        return None
    opis, v = rows[pos]
    if not re.search(label_rx, opis, re.I):
        return None
    return v


def parse_insurance(path: str) -> dict | None:
    """Osigurateljski obrazac (CROS): ISD (P&L, IFRS 17) + IFP (bilanca)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "ISD" not in wb.sheetnames or "IFP" not in wb.sheetnames:
        return None
    isd = _isd_rows(wb["ISD"])
    ifp = _isd_rows(wb["IFP"])
    if (_isd_get(isd, "001", r"Prihodi od ugovora o osiguranju") is None
            or _isd_get(ifp, "055", r"UKUPNA\s+AKTIVA") is None):
        return None
    items: dict[str, float] = {}
    src: dict[str, str] = {}

    def put(k, v, s):
        if v is not None:
            items[k] = v
            src[k] = s

    put("revenue", _isd_get(isd, "001", r"Prihodi od ugovora o osiguranju"),
        "ISD 001: prihodi od ugovora o osiguranju (IFRS 17; ista baza kao FY)")
    put("pretax_income", _isd_get(isd, "043", r"prije poreza"),
        "ISD 043: dobit prije poreza")
    put("net_income", _isd_get(isd, "047", r"poslije pore"),
        "ISD 047: dobit poslije poreza")
    put("net_income_parent", _isd_get(isd, "048", r"imateljima kapitala matice"),
        "ISD 048: pripisano matici")
    ni, nip = items.get("net_income"), items.get("net_income_parent")
    if ni is not None and nip is not None:
        put("net_income_minority", ni - nip, "ISD 047−048 (izračun)")
    put("total_assets", _isd_get(ifp, "055", r"UKUPNA\s+AKTIVA"),
        "IFP 055: ukupna aktiva")
    eq = _isd_get(ifp, "057", r"KAPITAL\s+I\s+REZERVE")
    mi = _isd_get(ifp, "078", r"MANJINSKI")
    put("total_equity", eq, "IFP 057: kapital i rezerve")
    if mi is not None:
        put("minority_interests", mi, "IFP 078: manjinski interes")
        if eq is not None:
            put("equity_parent", eq - mi, "IFP 057−078 (izračun)")
    return {"items": items, "src": src}


def parse_fin_any(path: str) -> tuple[dict | None, str]:
    """Proba tri layouta redom; -> (parsed, ime_layouta)."""
    p = parse_bank_tfi(path)
    if p and p["items"].get("total_assets"):
        return p, "nadzorni"
    p = parse_finrep(path)
    if p and p["items"].get("total_assets"):
        return p, "FINREP"
    p = parse_insurance(path)
    if p and p["items"].get("total_assets"):
        return p, "osigurateljski"
    return None, ""


def ingest_fin_interim(conn, ticker: str, url: str, year: int, period_type: str,
                       *, consolidated: bool = True, published_at=None) -> int | None:
    """Preuzmi + parsiraj interim financijske firme; vrati fid ili None."""
    os.makedirs(SCRATCH, exist_ok=True)
    path = f"{SCRATCH}/{ticker}_{year}_{period_type}.xlsx"
    r = requests.get(url.replace("\\/", "/"), timeout=120, verify=_verify())
    r.raise_for_status()
    with open(path, "wb") as fh:
        fh.write(r.content)
    parsed, layout = parse_fin_any(path)
    if not parsed:
        return None
    solo = "" if consolidated else " | SOLO obrazac"
    extraction = {
        "meta": {"company_ticker": ticker, "fiscal_year": year,
                 "period_type": period_type, "basis": "consolidated",
                 "audited": False, "cumulative": True,
                 "currency": "EUR", "reporting_scale": 1},
        "items": [
            {"item": k, "value_raw": v, "confidence": 0.9,
             "source_page": (f"{layout} obrazac {period_type} {year} XLSX "
                             f"(kumulativ), {parsed['src'][k]}{solo} — strojno "
                             "parsirano, nerevidirano")}
            for k, v in parsed["items"].items()],
    }
    return load_extraction(conn, extraction, source_url=url.replace("\\/", "/"),
                           doc_type="financial_report", published_at=published_at)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tickers", nargs="*", default=None)
    a = ap.parse_args(argv)
    today = datetime.date.today().isoformat()
    ok = review = skip = fail = 0
    with get_conn() as conn, conn.cursor() as cur:
        if a.tickers:
            firms = a.tickers
        else:
            cur.execute("""SELECT ticker FROM companies WHERE is_live
                           AND sector IN ('bank','insurance') ORDER BY ticker""")
            firms = [r[0] for r in cur.fetchall()]
        for tick in firms:
            try:
                d = feed("financialReports", ticker=tick,
                         date_from="2024-04-01", date_to=today)
                items = d.get("items") or []
            except Exception as e:  # noqa: BLE001
                print(f"[fin-interim] {tick}: feed pao ({e})"); fail += 1; continue
            best: dict = {}
            for it in items:
                if it.get("documentType") != "XLSX" or it.get("period") not in PERIOD_MAP:
                    continue
                key = (it["year"], PERIOD_MAP[it["period"]])
                if key not in best or (it["consolidated"] and not best[key]["consolidated"]):
                    best[key] = it
            n = 0
            for (year, pt), it in sorted(best.items()):
                try:
                    fid = ingest_fin_interim(
                        conn, tick, it["documentLink"], year, pt,
                        consolidated=bool(it.get("consolidated")),
                        published_at=it.get("publishDate"))
                    if fid is None:
                        print(f"[fin-interim] {tick} {year}{pt}: nepoznat layout -> preskačem")
                        skip += 1
                        continue
                    conn.commit()
                    res = validate_filing(conn, fid)
                    if res["status"] == "validated":
                        ok += 1
                    else:
                        review += 1
                    n += 1
                except Exception as e:  # noqa: BLE001
                    conn.rollback()
                    print(f"[fin-interim] {tick} {year}{pt}: GREŠKA {str(e)[:80]}")
                    fail += 1
            print(f"[fin-interim] {tick}: {n} perioda upisano")
    print(f"\nGOTOVO: validated={ok}, needs_review={review}, skip={skip}, fail={fail}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
