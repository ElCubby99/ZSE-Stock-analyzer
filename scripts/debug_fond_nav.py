#!/usr/bin/env python3
"""Jednokratni forenzički alat (runda 4): REGOS statistički pokazatelji.

Runda 3: regos.hr dostupan; stranica 'statisticki-pokazatelji-za-2026-godinu'
postoji ali nije dumpana. Runda 4: dump te stranice (svi linkovi + datoteke),
preuzmi najnovije XLSX/PDF kandidate i ispiši strukturu (sheetovi + redovi)
— tražimo NETO IMOVINU PO POJEDINOM OMF-u. Uz to: hrportfolio tablica
(orijentir gdje agregatori nalaze per-fond NAV).
"""
import io
import re
import sys

import requests

UA = {"User-Agent": "Mozilla/5.0 (compatible; podaci; burzovnilist.com)"}


def _fetch(url, timeout=45):
    r = requests.get(url, timeout=timeout, headers=UA)
    r.raise_for_status()
    return r


def main() -> int:
    print("########## REGOS statistički pokazatelji ##########")
    for url in ("https://regos.hr/statisticki-pokazatelji/statisticki-pokazatelji-za-2026-godinu/",
                "https://regos.hr/statisticki-pokazatelji/"):
        try:
            html = _fetch(url).text
        except Exception as e:  # noqa: BLE001
            print(f"  {url}: GREŠKA {type(e).__name__}: {str(e)[:100]}")
            continue
        print(f"  --- {url} ---")
        files = []
        for m in re.finditer(r'<a\b[^>]*href="([^"]+)"[^>]*>(.*?)</a>', html, re.I | re.S):
            href = m.group(1)
            txt = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", m.group(2))).strip()
            if re.search(r"\.(xlsx?|pdf|csv)($|\?)", href, re.I):
                u = href if href.startswith("http") else "https://regos.hr/" + href.lstrip("/")
                files.append((u, txt))
                print(f"    FILE: {u[:150]}  TEXT: {txt[:80]}")
            elif re.search(r"pokazatelj|statisti|imovin|fond", f"{href} {txt}", re.I):
                print(f"    LINK: {href[:130]}  TEXT: {txt[:80]}")
        # preuzmi do 3 kandidata i ispiši strukturu
        for u, txt in files[:3]:
            print(f"  === STRUKTURA: {u} ({txt[:60]}) ===")
            try:
                raw = _fetch(u, 60).content
                if raw.startswith(b"PK"):
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                    for ws in wb.worksheets:
                        print(f"    --- sheet: {ws.title!r} ---")
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i >= 14:
                                break
                            cells = [str(c)[:30] if c is not None else "" for c in row[:9]]
                            print("     |", " | ".join(cells))
                elif raw[:5] == b"%PDF-":
                    print(f"    PDF ({len(raw)} B) — tekstualni sloj:")
                    txt2 = raw.decode("latin-1", errors="ignore")
                    # gruba ekstrakcija teksta iz PDF streamova nije pouzdana —
                    # samo signal ima li 'imovina' u raw sadržaju
                    hits = len(re.findall(r"[Ii]movin", txt2))
                    print(f"    'imovin' u raw PDF: {hits}")
                else:
                    print(f"    nepoznat format: {raw[:20]!r}")
            except Exception as e:  # noqa: BLE001
                print(f"    GREŠKA {type(e).__name__}: {str(e)[:110]}")
        break

    print("\n########## hrportfolio tablica (orijentir) ##########")
    try:
        html = _fetch("https://hrportfolio.hr/fondovi/pregled-mirovinskih-fondova").text
        for i, tm in enumerate(re.finditer(r"<table.*?</table>", html, re.I | re.S)):
            if i >= 2:
                break
            rows = re.findall(r"<tr.*?</tr>", tm.group(0), re.I | re.S)
            print(f"  TABLICA {i + 1} ({len(rows)} redova):")
            for r_ in rows[:16]:
                cells = [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", c)).strip()[:34]
                         for c in re.findall(r"<t[dh].*?</t[dh]>", r_, re.I | re.S)]
                print("   |", " | ".join(cells[:8]))
    except Exception as e:  # noqa: BLE001
        print(f"  GREŠKA {type(e).__name__}: {str(e)[:100]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
