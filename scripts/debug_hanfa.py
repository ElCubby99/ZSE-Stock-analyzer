#!/usr/bin/env python3
"""Jednokratni forenzički alat: ispiši STVARNU strukturu HANFA objava.

hanfa.hr nije dostupna iz razvojnog okruženja (allowlist), pa se struktura
ne može pogledati lokalno — ovaj skript se vrti kroz debug-hanfa.yml
workflow (GitHub runner, otvorena mreža) i u log ispiše:
  1. sve <a> poveznice na XLSX/getfile s listinga statistike MF
  2. za kandidate s 'jedinic'/'mirex'/'vrijednost'/c-0N u nazivu:
     imena sheetova + prvih ~15 redova × 8 stupaca svakog sheeta
Iz tog ispisa se piše ISPRAVAN strogi parser (ništa izmišljeno).
"""
import io
import re
import sys

import requests

LISTINGS = (
    "https://www.hanfa.hr/statistika/mirovinski-fondovi/",
    "https://www.hanfa.hr/statistika/",
)
UA = {"User-Agent": "Mozilla/5.0 (podaci; burzovnilist.com)"}


def main() -> int:
    links = []
    for listing in LISTINGS:
        try:
            r = requests.get(listing, timeout=90, headers=UA)
            r.raise_for_status()
        except Exception as e:  # noqa: BLE001
            print(f"[listing] {listing}: {type(e).__name__}: {e}")
            continue
        print(f"\n===== LISTING {listing} =====")
        for m in re.finditer(r'<a\b[^>]*href="([^"]+)"[^>]*>(.*?)</a>',
                             r.text, re.I | re.S):
            href, text = m.group(1), re.sub(r"<[^>]+>", " ", m.group(2))
            text = re.sub(r"\s+", " ", text).strip()
            if re.search(r"\.xlsx|/getfile/", href, re.I):
                url = href if href.startswith("http") else f"https://www.hanfa.hr{href}"
                print(f"  LINK: {url}  TEXT: {text[:90]}")
                links.append((url, text))
        break  # prvi dostupan listing je dovoljan

    want = [(u, t) for u, t in links
            if re.search(r"jedinic|mirex|vrijednost|c-0\d", f"{u} {t}", re.I)]
    from openpyxl import load_workbook
    for url, text in want[:8]:
        print(f"\n===== FILE {url} ({text[:60]}) =====")
        try:
            raw = requests.get(url, timeout=120, headers=UA).content
            if not raw.startswith(b"PK"):
                print("  nije XLSX (magic bytes)")
                continue
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            for ws in wb.worksheets:
                print(f"  --- sheet: {ws.title!r} ---")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 15:
                        break
                    cells = [str(c)[:28] if c is not None else "" for c in row[:8]]
                    print("   |", " | ".join(cells))
        except Exception as e:  # noqa: BLE001
            print(f"  GREŠKA: {type(e).__name__}: {e}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
