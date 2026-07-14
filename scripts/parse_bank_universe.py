"""M19-B: deterministički parser NADZORNOG obrasca za banke (EHO XLSX) — 0 kredita.

Male banke (IKBA, KBZ, PDBA, SNBA) objavljuju godišnji FI kao standardizirani
nadzorni XLSX obrazac (drugi AOP-ovi od TFI-POD-a). Ovaj parser mapira taj
obrazac na bankovnu taksonomiju (canonical BANK_INCOME/BANK_BALANCE) po
STABILNIM AOP pozicijama, uz obavezno poklapanje labela (sigurnost od drifta
obrasca — bez poklapanja se stavka NE puni; ništa izmišljeno).

Mapiranje (RDG / Bilanca nadzornog obrasca, kumulativ tekuće godine):
  net_interest_income  = AOP1 (kamatni prihodi) − AOP3 (kamatni rashodi)
  net_fee_income       = AOP8 (prihodi od provizija) − AOP9 (rashodi za provizije)
  total_operating_income = NII + net naknade + AOP4 (VP) + AOP10 (fin. aktivnosti)
                           + AOP11 (ostali prihodi iz redovnog poslovanja)
  loan_loss_provisions = AOP18 (rezerviranja) + AOP21 (umanjenja krediti)
                         + AOP22 (umanjenja VP)  [predznak kako je objavljen]
  pretax_income = AOP23 + AOP26; income_tax = AOP24 + AOP27;
  net_income    = AOP25 + AOP28 (nastavljeno + obustavljeno poslovanje)
  Bilanca: total_assets=AOP25, cash=AOP1, loans_to_customers=AOP8,
  deposits_from_customers=AOP29, total_equity = UKUPNE OBVEZE I KAPITAL −
  zbroj obveza (identitet), manjinski = AOP62.

Pokretanje:  python -m scripts.parse_bank_universe [--tickers IKBA KBZ PDBA SNBA]
"""
from __future__ import annotations

import argparse
import datetime
import os
import re
import sys

sys.path.insert(0, ".")

import openpyxl  # noqa: E402
import requests  # noqa: E402

from scripts.parse_tfi_universe import _verify  # noqa: E402
from src.db import get_conn  # noqa: E402
from src.eho import feed  # noqa: E402
from src.loader import load_extraction  # noqa: E402
from src.onboard import core_gate  # noqa: E402
from src.validator import validate_filing  # noqa: E402

SCRATCH = "/tmp/bank_xlsx"


def pick_bank_report(items: list[dict]) -> dict | None:
    """SAMO 1Y XLSX (godišnji nadzorni obrazac — verificirani layout);
    4Q kvartalni je FINREP varijanta s drugim AOP-ovima -> zaseban parser."""
    cands = [x for x in items
             if x.get("documentType") == "XLSX" and x.get("period") == "1Y"
             and (x.get("year") or 0) >= 2024]
    if not cands:
        return None
    cands.sort(key=lambda x: (x["year"], bool(x["consolidated"])), reverse=True)
    return cands[0]


def _aop_rows(ws) -> dict[int, tuple[str, float | None]]:
    """{aop: (label, tekuća_vrijednost)} — AOP + zadnja numerička kolona."""
    out: dict[int, tuple[str, float | None]] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=11):
        vals = [c.value for c in row]
        label = next((str(v).strip() for v in vals if isinstance(v, str) and v.strip()), None)
        if not label:
            continue
        nums = [(i, v) for i, v in enumerate(vals) if isinstance(v, (int, float))]
        if not nums:
            continue
        aop_i, aop_v = nums[0]
        if not (0 < aop_v < 400 and float(aop_v).is_integer()):
            continue
        tail = [v for i, v in nums[1:]]
        # nadzorni 1Y: [prior, tekuća] -> zadnja; kvartalni s 4 kolone -> 3.
        cur = (tail[2] if len(tail) >= 4 else tail[-1] if tail else None)
        aop = int(aop_v)
        if aop not in out:  # prvi red s tim AOP-om (RDG i Bilanca se parsiraju odvojeno)
            out[aop] = (label, float(cur) if cur is not None else None)
    return out


def _get(rows: dict, aop: int, label_rx: str) -> float | None:
    """Vrijednost SAMO ako se i AOP i labela poklapaju (zaštita od drifta)."""
    if aop not in rows:
        return None
    label, val = rows[aop]
    if not re.search(label_rx, label, re.I):
        return None
    return val


def parse_bank_tfi(path: str) -> dict | None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "RDG" not in wb.sheetnames or "Bilanca" not in wb.sheetnames:
        return None
    rdg = _aop_rows(wb["RDG"])
    bil = _aop_rows(wb["Bilanca"])

    # sanity: je li ovo nadzorni obrazac? (TFI-POD ima 'POSLOVNI PRIHODI')
    if _get(rdg, 1, r"kamat") is None or _get(bil, 25, r"UKUPNA IMOVINA") is None:
        return None

    items: dict[str, float] = {}
    src: dict[str, str] = {}

    def put(k, v, s):
        if v is not None:
            items[k] = v
            src[k] = s

    ii = _get(rdg, 1, r"kamata i slični prihodi")
    ie = _get(rdg, 3, r"kamata i slični rashodi")
    if ii is not None and ie is not None:
        put("net_interest_income", ii - ie, "RDG AOP 001−003 (kamatni prihodi − rashodi)")
    fi = _get(rdg, 8, r"provizija")
    fe = _get(rdg, 9, r"provizije")
    if fi is not None and fe is not None:
        put("net_fee_income", fi - fe, "RDG AOP 008−009 (provizije neto)")
    vp = _get(rdg, 4, r"vrijednosnih papira") or 0.0
    fin = _get(rdg, 10, r"financijskih aktivnosti") or 0.0
    ost = _get(rdg, 11, r"redovnog poslovanja") or 0.0
    if "net_interest_income" in items and "net_fee_income" in items:
        put("total_operating_income",
            items["net_interest_income"] + items["net_fee_income"] + vp + fin + ost,
            "RDG: NII + neto provizije + AOP 004 + 010 + 011 (izračun)")
    rez = _get(rdg, 18, r"[Rr]ezerviranja") or 0.0
    umk = _get(rdg, 21, r"kredita") or 0.0
    umv = _get(rdg, 22, r"vrijednosn") or 0.0
    put("loan_loss_provisions", rez + umk + umv,
        "RDG AOP 018+021+022 (rezerviranja + umanjenja; + = trošak)")
    pt1 = _get(rdg, 23, r"prije oporezivanja")
    pt2 = _get(rdg, 26, r"prije oporezivanja") or 0.0
    if pt1 is not None:
        put("pretax_income", pt1 + pt2, "RDG AOP 023(+026)")
    tx1 = _get(rdg, 24, r"[Pp]orez")
    tx2 = _get(rdg, 27, r"[Pp]orez") or 0.0
    if tx1 is not None:
        put("income_tax", tx1 + tx2, "RDG AOP 024(+027)")
    ni1 = _get(rdg, 25, r"nakon oporezivanja")
    ni2 = _get(rdg, 28, r"nakon oporezivanja") or 0.0
    if ni1 is not None:
        put("net_income", ni1 + ni2, "RDG AOP 025(+028)")

    put("total_assets", _get(bil, 25, r"UKUPNA IMOVINA"), "Bilanca AOP 025")
    put("cash_and_equivalents", _get(bil, 1, r"[Nn]ovac"),
        "Bilanca AOP 001 (novac + središnje banke)")
    put("loans_to_customers", _get(bil, 8, r"[Kk]rediti i predujmovi klijentima"),
        "Bilanca AOP 008")
    put("deposits_from_customers", _get(bil, 29, r"Obveze prema klijentima"),
        "Bilanca AOP 029")
    mi = _get(bil, 62, r"Manjinski")
    if mi is not None:
        put("minority_interests", mi, "Bilanca AOP 062")
    # kapital identitetom: UKUPNE OBVEZE I KAPITAL − sve obveze (AOP 26..46)
    total = _get(bil, 63, r"UKUPNE OBVEZE I KAPITAL")
    liab_aops = [26, 29, 36, 37, 38, 39, 42, 45, 46]
    liabs = [rows_v for a in liab_aops
             if (rows_v := (bil.get(a, (None, None))[1])) is not None]
    if total is not None and len(liabs) == len(liab_aops):
        eq = total - sum(liabs)
        put("total_equity", eq,
            "Bilanca: identitet — AOP 063 − obveze (026,029,036..046)")
        if mi is not None:
            put("equity_parent", eq - mi, "Bilanca: kapital − manjinski (identitet)")
    ni = items.get("net_income")
    if ni is not None and (mi or 0) == 0:
        put("net_income_parent", ni, "RDG: nema manjinskih (AOP 062 = 0) — sve matici")
        put("net_income_minority", 0.0, "Bilanca AOP 062 = 0")
    # broj zaposlenih iz 'Opći podaci'
    if "Opći podaci" in wb.sheetnames:
        for row in wb["Opći podaci"].iter_rows(max_col=10):
            vals = [c.value for c in row]
            if any(isinstance(v, str) and re.search(r"Broj zaposlenih", v) for v in vals):
                n = next((v for v in vals if isinstance(v, (int, float)) and v > 0), None)
                if n:
                    put("employees", float(n), "Opći podaci: broj zaposlenih")
                break
    return {"items": items, "src": src}


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tickers", nargs="*", default=["IKBA", "KBZ", "PDBA", "SNBA"])
    a = ap.parse_args(argv)
    os.makedirs(SCRATCH, exist_ok=True)
    today = datetime.date.today().isoformat()
    promoted, review, skipped = [], [], []
    with get_conn() as conn, conn.cursor() as cur:
        for tick in a.tickers:
            try:
                d = feed("financialReports", ticker=tick,
                         date_from="2025-01-01", date_to=today)
                rep = pick_bank_report(d.get("items") or [])
                if not rep:
                    skipped.append(f"{tick} (nema 1Y XLSX na EHO)")
                    continue
                path = f"{SCRATCH}/{tick}.xlsx"
                r = requests.get(rep["documentLink"].replace("\\/", "/"),
                                 timeout=120, verify=_verify())
                r.raise_for_status()
                with open(path, "wb") as fh:
                    fh.write(r.content)
                parsed = parse_bank_tfi(path)
                if not parsed or not parsed["items"].get("total_assets"):
                    skipped.append(f"{tick} (obrazac nije nadzorni ili prazan)")
                    continue
                solo = ("" if rep["consolidated"] else
                        " | SOLO obrazac (banka bez konsolidiranog)")
                extraction = {
                    "meta": {"company_ticker": tick, "fiscal_year": rep["year"],
                             "period_type": "annual", "basis": "consolidated",
                             "audited": False, "cumulative": True,
                             "currency": "EUR", "reporting_scale": 1},
                    "items": [
                        {"item": k, "value_raw": v, "confidence": 0.9,
                         "source_page": (f"Nadzorni obrazac {rep['period']} "
                                         f"{rep['year']} XLSX, {parsed['src'][k]}{solo}"
                                         " — strojno parsirano, nerevidirano")}
                        for k, v in parsed["items"].items()],
                }
                fid = load_extraction(conn, extraction,
                                      source_url=rep["documentLink"].replace("\\/", "/"),
                                      doc_type="financial_report",
                                      published_at=rep.get("publishDate"))
                conn.commit()
                res = validate_filing(conn, fid)
                cur.execute("SELECT company_id FROM filings WHERE id=%s", (fid,))
                cid = cur.fetchone()[0]
                blocking, _notes = core_gate(conn, fid, "bank")
                if res["status"] == "validated" and not blocking:
                    cur.execute("""UPDATE companies SET is_live=TRUE,
                                   onboarding_status='live' WHERE id=%s""", (cid,))
                    conn.commit()
                    promoted.append(tick)
                    print(f"[bank] {tick}: filing {fid} VALIDATED, gate čist -> LIVE "
                          f"({len(parsed['items'])} stavki)")
                else:
                    fails = [x["detail"][:80] for x in res["results"]
                             if x["status"] in ("FAIL", "WARN")] + blocking
                    review.append(f"{tick}: {fails}")
                    print(f"[bank] {tick}: filing {fid} {res['status']} — {fails}")
                    conn.commit()
            except Exception as e:  # noqa: BLE001
                conn.rollback()
                skipped.append(f"{tick} (GREŠKA: {str(e)[:70]})")
    print(f"\nPROMOVIRANO ({len(promoted)}): {' '.join(promoted)}")
    print(f"NEEDS_REVIEW ({len(review)}):")
    for x in review:
        print(f"  - {x}")
    print(f"PRESKOČENO ({len(skipped)}):")
    for x in skipped:
        print(f"  - {x}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
