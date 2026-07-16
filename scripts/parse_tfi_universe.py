"""M16-B: masovna ekstrakcija financija CIJELOG univerzuma iz TFI XLSX — bez API.

Izvor: EHO financialReports feed -> standardizirani TFI-POD XLSX obrazac
(Bilanca/RDG/NT_I s AOP oznakama). Za svaku 'discovered' firmu uzima se
NAJNOVIJI 4Q (kumulativ 1.1.–31.12. = FY, nerevidirano) ili 1Y XLSX,
konsolidirani ako postoji, inače solo (izdavatelj bez grupe — označeno).
Parsirane stavke idu kroz POSTOJEĆI put: loader -> validator -> core gate;
promocija u live SAMO kad je sve čisto. Banke/osiguratelji imaju drugačiji
obrazac -> v1 ih preskače uz napomenu (ništa se ne izmišlja).

Pokretanje:  python -m scripts.parse_tfi_universe [--tickers T1 T2 ...]
"""
from __future__ import annotations

import argparse
import datetime
import os
import re
import sys

sys.path.insert(0, ".")

import requests  # noqa: E402

from src.db import get_conn  # noqa: E402
from src.eho import feed  # noqa: E402
from src.loader import load_extraction  # noqa: E402
from src.onboard import core_gate  # noqa: E402
from src.validator import validate_filing  # noqa: E402

SCRATCH = "/tmp/tfi_xlsx"

# EHO period oznaka <-> naš period_type. Interim (IFRS) je KUMULATIV (YTD).
LABEL_TO_PT = {"1Y": "annual", "1Q": "q1", "2Q": "h1", "3Q": "9m", "4Q": "q4"}
PT_TO_LABEL = {v: k for k, v in LABEL_TO_PT.items()}


def _verify():
    return os.getenv("REQUESTS_CA_BUNDLE") or os.getenv("SSL_CERT_FILE") or True


def ingest_tfi_xlsx(conn, ticker, url, year, period_type, *, consolidated=True,
                    cumulative=True, published_at=None, path=None,
                    currency="EUR"):
    """Preuzmi + parsiraj TFI-POD XLSX i upiši kao filing (deterministički,
    0 kredita). Vraća (fid, parsed) ili (None, None) ako nije TFI-POD obrazac
    (npr. bankovni nadzorni — traži zaseban parser). Ne commita."""
    import requests
    if path is None:
        os.makedirs(SCRATCH, exist_ok=True)
        path = f"{SCRATCH}/{ticker}_{year}_{period_type}.xlsx"
    r = requests.get(url, timeout=120, verify=_verify())
    r.raise_for_status()
    with open(path, "wb") as fh:
        fh.write(r.content)
    parsed = parse_tfi(path)
    if not parsed or not parsed["items"]:
        return None, None
    solo = "" if consolidated else " | SOLO obrazac (izdavatelj bez konsolidiranog TFI-ja)"
    label = PT_TO_LABEL.get(period_type, period_type)
    kind = "kumulativ" if cumulative else "iznos"
    extraction = {
        "meta": {"company_ticker": ticker, "fiscal_year": year,
                 "period_type": period_type, "basis": "consolidated",
                 "audited": False, "cumulative": cumulative,
                 # M35: izvještaji do FY2022 su u HRK — normalize dijeli
                 # fiksnim tečajem (7,5345); od FY2023 EUR
                 "currency": currency, "reporting_scale": 1},
        "items": [
            {"item": k, "value_raw": v, "confidence": 0.9 if k != "ebit" else 0.85,
             "source_page": (f"TFI {label} {year} XLSX ({kind}), {parsed['src'][k]}"
                             f"{solo} — strojno parsirano (AOP obrazac), nerevidirano")}
            for k, v in parsed["items"].items()],
    }
    fid = load_extraction(conn, extraction, source_url=url,
                          doc_type="financial_report", published_at=published_at)
    return fid, parsed


def _num(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(".", "").replace(",", ".")
        if not v or v in ("-", "–"):
            return None
        try:
            v = float(v)
        except ValueError:
            return None
    return float(v)


def _rows(ws):
    """-> [(label, aop, prior, current)] — AOP stupac + zadnja dva numerička."""
    out = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=11):
        vals = [c.value for c in row]
        label = next((str(v).strip() for v in vals if isinstance(v, str) and v.strip()), None)
        if not label:
            continue
        nums = [(i, v) for i, v in enumerate(vals[1:], start=1) if v is not None]
        aop = None
        for i, v in nums:
            if isinstance(v, (int, float)) and 0 < v < 400 and float(v).is_integer() \
                    and (vals[i + 1] is not None or vals[i - 1] is not None or True):
                aop = int(v)
                tail = [_num(x) for j, x in nums if j > i]
                prior = tail[0] if tail else None
                # RDG kvartalni obrazac ima 4 stupca (prior kumulativ, prior
                # tromjesečje, TEKUĆI KUMULATIV, tekuće tromjesečje) -> kumulativ
                # je 3. stupac; bilanca/NT imaju 2 -> zadnji
                cur = (tail[2] if len(tail) >= 4 else
                       tail[-1] if len(tail) >= 2 else None)
                out.append((label, aop, prior, cur))
                break
    return out


def _find(rows, pattern, aop=None):
    rx = re.compile(pattern, re.I)
    for label, a, prior, cur in rows:
        if aop is not None and a != aop:
            continue
        if rx.search(label):
            return cur
    return None


def _find_in_section(rows, start_rx, end_rx, item_rx, agg=False):
    """Vrijednost (ili zbroj uz agg=True) retka unutar sekcije obrasca."""
    inside = False
    total, hit = 0.0, False
    srx, erx, irx = re.compile(start_rx, re.I), re.compile(end_rx), re.compile(item_rx, re.I)
    for label, a, prior, cur in rows:
        if srx.search(label):
            inside = True
            continue
        if inside and erx.match(label):
            break
        if inside and irx.search(label) and cur is not None:
            if not agg:
                return cur
            total += cur
            hit = True
    return (total if hit else None) if agg else None


def parse_tfi(path: str) -> dict | None:
    """XLSX -> {item: value_eur} (tekuće razdoblje/kumulativ) ili None."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    bil_name = next((n for n in ("Bilanca", "BS", "BIL") if n in wb.sheetnames), None)
    if not bil_name or "RDG" not in wb.sheetnames:
        return None  # drugačiji obrazac (banka/osiguratelj/fond)
    bil = _rows(wb[bil_name])
    rdg = _rows(wb["RDG"])
    items: dict = {}
    src: dict = {}

    def put(item, val, where):
        if val is not None:
            items[item] = val
            src[item] = where

    put("total_assets", _find(bil, r"UKUPNO\s+AKTIVA"), "Bilanca: UKUPNO AKTIVA")
    put("total_equity",
        _find(bil, r"^A\)\s*KAPITAL I REZERVE") or _find(bil, r"^A\s+KAPITAL I REZERVE"),
        "Bilanca: KAPITAL I REZERVE")
    put("equity_parent", _find(bil, r"PRIPISANO IMATELJIMA KAPITALA MATICE"),
        "Bilanca: pripisano imateljima kapitala matice")
    put("minority_interests", _find(bil, r"MANJINSKI|NEKONTROLIRAJUĆI"),
        "Bilanca: manjinski interes")
    put("cash_and_equivalents",
        _find(bil, r"NOVAC U BANCI I BLAGAJNI")
        or _find(bil, r"^III\s+NOVAC I NOVČANI EKVIVALENTI"),
        "Bilanca: novac u banci i blagajni")
    # kamatonosni dug: banke+krediti+obveznice u dugoročnim (C) i kratkoročnim (D)
    sec = None
    dl = ds = None
    for label, a, prior, cur in bil:
        if re.match(r"^C\)\s*DUGOROČNE OBVEZE", label, re.I):
            sec = "L"
        elif re.match(r"^D\)\s*KRATKOROČNE OBVEZE", label, re.I):
            sec = "S"
        elif re.match(r"^[E-J]\)", label):
            sec = None
        if sec and re.search(r"banka|bankama|vrijednosnim papirima|zajmove", label, re.I) \
                and re.search(r"obveze", label, re.I) and cur is not None:
            if sec == "L":
                dl = (dl or 0) + cur
            else:
                ds = (ds or 0) + cur
    put("debt_long", dl, "Bilanca: dugoročne obveze prema bankama/za zajmove/po vp")
    put("debt_short", ds, "Bilanca: kratkoročne obveze prema bankama/za zajmove/po vp")
    # M18: proširena taksonomija (tekući omjer, EV, DSO/DIO/DPO, Altman)
    put("current_assets", _find(bil, r"^C\)\s*KRATKOTRAJNA IMOVINA"),
        "Bilanca: C) KRATKOTRAJNA IMOVINA")
    put("current_liabilities", _find(bil, r"^D\)\s*KRATKOROČNE OBVEZE"),
        "Bilanca: D) KRATKOROČNE OBVEZE")
    put("short_term_fin_assets", _find(bil, r"^III\.\s*KRATKOTRAJNA FINANCIJSKA"),
        "Bilanca: III. kratkotrajna financijska imovina")
    put("retained_earnings", _find(bil, r"^\s*1?\.?\s*Zadržana dobit"),
        "Bilanca: zadržana dobit")
    put("inventories", _find(bil, r"^I\.\s*ZALIHE"), "Bilanca: I. zalihe")
    put("trade_receivables",
        _find_in_section(bil, r"^C\)\s*KRATKOTRAJNA IMOVINA", r"^[D-J]\)",
                         r"Potraživanja od kupaca"),
        "Bilanca: potraživanja od kupaca (kratkotrajna)")
    put("trade_payables",
        _find_in_section(bil, r"^D\)\s*KRATKOROČNE OBVEZE", r"^[E-J]\)",
                         r"Obveze prema dobavljačima"),
        "Bilanca: obveze prema dobavljačima (kratkoročne)")

    rev = _find(rdg, r"^I\.\s*POSLOVNI PRIHODI")
    opex = _find(rdg, r"^II\.\s*POSLOVNI RASHODI")
    # Varijanta obrasca za financijske usluge (npr. ZB — burza): sekcije su
    # slovima A–J umjesto rimskim brojevima; iste veličine, drugi anchori.
    if rev is None:
        rev = _find(rdg, r"^A\s+POSLOVNI PRIHODI")
    if opex is None:
        opex = _find(rdg, r"^B\s+POSLOVNI RASHODI")
    put("revenue", rev, "RDG: poslovni prihodi (kumulativ)")
    put("operating_expenses", opex, "RDG: poslovni rashodi (kumulativ)")
    put("depreciation_amortization",
        _find(rdg, r"^\s*4?\.?\s*Amortizacija") or _find(rdg, r"^III\s+Amortizacija"),
        "RDG: amortizacija")
    put("material_costs", _find(rdg, r"Materijalni troškovi"),
        "RDG: materijalni troškovi (COGS proxy)")
    put("interest_expense",
        _find_in_section(rdg, r"^IV\.\s*FINANCIJSKI RASHODI", r"^(V|X)I*\.",
                         r"kamat", agg=True)
        or _find_in_section(rdg, r"^D\s+FINANCIJSKI RASHODI", r"^[E-J]\s+",
                            r"kamat", agg=True),
        "RDG: rashodi od kamata (unutar financijskih rashoda, zbroj)")
    if rev is not None and opex is not None:
        put("ebit", rev - opex, "RDG: poslovni prihodi − poslovni rashodi (izračun)")
    fp = _find(rdg, r"^III\.\s*FINANCIJSKI PRIHODI")
    fr_ = _find(rdg, r"^IV\.\s*FINANCIJSKI RASHODI")
    if fp is None:
        fp = _find(rdg, r"^C\s+FINANCIJSKI PRIHODI")
    if fr_ is None:
        fr_ = _find(rdg, r"^D\s+FINANCIJSKI RASHODI")
    if fp is not None and fr_ is not None:
        put("net_financial_result", fp - fr_, "RDG: fin. prihodi − fin. rashodi")
    put("pretax_income", _find(rdg, r"DOBIT ILI GUBITAK PRIJE OPOREZIVANJA"),
        "RDG: dobit prije oporezivanja")
    put("income_tax",
        _find(rdg, r"^\s*XII?I?\.\s*POREZ NA DOBIT|^POREZ NA DOBIT")
        or _find(rdg, r"^I\s+POREZ NA DOBIT"),
        "RDG: porez na dobit")
    put("net_income", _find(rdg, r"DOBIT ILI GUBITAK RAZDOBLJA(?!.*(matice|manjinsk))"),
        "RDG: dobit ili gubitak razdoblja")
    put("net_income_parent", _find(rdg, r"Pripisana imateljima kapitala matice"),
        "RDG: pripisano matici")
    put("net_income_minority", _find(rdg, r"Pripisana manjinskom"),
        "RDG: pripisano manjinskom interesu")
    # SOLO obrazac: retci matica/manjine postoje s 0 — dobit pripada u
    # cijelosti izdavatelju (nema manjinskih), pa se 0-retci preslikavaju
    ni = items.get("net_income")
    if ni and not items.get("net_income_parent") and not items.get("net_income_minority"):
        put("net_income_parent", ni, "RDG: solo obrazac — cijela dobit izdavatelju")
        put("net_income_minority", 0.0, "RDG: solo obrazac — nema manjinskih")
    eq = items.get("total_equity")
    if eq and not items.get("equity_parent"):
        mi = items.get("minority_interests")
        if mi:
            put("equity_parent", eq - mi,
                "Bilanca: identitet — KAPITAL I REZERVE − manjinski interes "
                "(obrazac bez retka 'pripisano matici')")
        else:
            put("equity_parent", eq, "Bilanca: solo obrazac — sav kapital izdavatelju")
    if "NT_I" in wb.sheetnames:
        nt = _rows(wb["NT_I"])
        put("operating_cf", _find(nt, r"NETO NOVČANI TOKOVI OD POSLOVNIH"),
            "NT_I: neto novčani tokovi od poslovnih aktivnosti")
        capex = _find(nt, r"kupnju dugotrajne")
        if capex is not None:
            put("capex", abs(capex), "NT_I: izdaci za kupnju dugotrajne imovine")
        put("investing_cf", _find(nt, r"NETO NOVČANI TOKOVI OD INVESTICIJSKIH"),
            "NT_I: B) neto tokovi od investicijskih aktivnosti")
        put("financing_cf", _find(nt, r"NETO NOVČANI TOKOVI OD FINANCIJSKIH"),
            "NT_I: C) neto tokovi od financijskih aktivnosti")
        # Starija NT varijanta (npr. ZB): nema NETO retke — neto tok = Ukupno
        # povećanje − Ukupno smanjenje / primici − izdaci (objavljeni retci)
        if items.get("operating_cf") is None:
            up = _find(nt, r"Ukupno povećanje novčanog tijeka od poslovnih")
            dn = _find(nt, r"Ukupno smanjenje novčanog tijeka od poslovnih")
            if up is not None and dn is not None:
                put("operating_cf", up - dn,
                    "NT_I: povećanje − smanjenje od poslovnih aktivnosti (izračun)")
        if items.get("investing_cf") is None:
            up = _find(nt, r"Ukupno novčani primici od investicijskih")
            dn = _find(nt, r"Ukupno novčani izdaci od investicijskih")
            if up is not None and dn is not None:
                put("investing_cf", up - dn,
                    "NT_I: primici − izdaci od investicijskih aktivnosti (izračun)")
        if items.get("financing_cf") is None:
            up = _find(nt, r"Ukupno novčani primici od financijskih")
            dn = _find(nt, r"Ukupno novčani izdaci od financijskih")
            if up is not None and dn is not None:
                put("financing_cf", up - dn,
                    "NT_I: primici − izdaci od financijskih aktivnosti (izračun)")
    # M20: firme koje NT izvještavaju DIREKTNOM metodom (NT_D) imaju NT_I list
    # pun nula — nula na neto tokovima je parse-fail signatura, ne stvarna
    # vrijednost. Tada se čita NT_D (isti NETO retci; izdaci s minusom).
    if not items.get("operating_cf") and "NT_D" in wb.sheetnames:
        ntd = _rows(wb["NT_D"])
        ocf_d = _find(ntd, r"NETO NOVČANI TOKOVI OD POSLOVNIH")
        if ocf_d:
            put("operating_cf", ocf_d,
                "NT_D (direktna metoda): neto novčani tokovi od poslovnih aktivnosti")
            capex_d = _find(ntd, r"kupnju dugotrajne")
            if capex_d:
                put("capex", abs(capex_d), "NT_D: izdaci za kupnju dugotrajne imovine")
            inv_d = _find(ntd, r"NETO NOVČANI TOKOVI OD INVESTICIJSKIH")
            if inv_d is not None:
                put("investing_cf", inv_d, "NT_D: B) neto tokovi od investicijskih")
            fin_d = _find(ntd, r"NETO NOVČANI TOKOVI OD FINANCIJSKIH")
            if fin_d is not None:
                put("financing_cf", fin_d, "NT_D: C) neto tokovi od financijskih")
    # M18: broj zaposlenih iz 'Opći podaci' (count, bez skale)
    if "Opći podaci" in wb.sheetnames:
        for row in wb["Opći podaci"].iter_rows(max_col=10):
            vals = [c.value for c in row]
            if any(isinstance(v, str) and re.search(r"Broj zaposlenih", v) for v in vals):
                n = next((v for v in vals if isinstance(v, (int, float)) and v > 0), None)
                if n:
                    put("employees", float(n), "Opći podaci: broj zaposlenih (kraj razdoblja)")
                break
    return {"items": items, "src": src}


def pick_report(items: list[dict]) -> dict | None:
    """Najnoviji FY: 4Q (kumulativ) ili 1Y, XLSX; konsolidirani ima prednost."""
    cands = [x for x in items
             if x.get("documentType") == "XLSX" and x.get("period") in ("4Q", "1Y")
             and (x.get("year") or 0) >= 2024]
    if not cands:
        return None
    cands.sort(key=lambda x: (x["year"], x["consolidated"], x["period"] == "4Q"),
               reverse=True)
    return cands[0]


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tickers", nargs="*", default=None)
    a = ap.parse_args(argv)
    os.makedirs(SCRATCH, exist_ok=True)
    today = datetime.date.today().isoformat()
    promoted, review, skipped = [], [], []
    with get_conn() as conn, conn.cursor() as cur:
        if a.tickers:
            todo = a.tickers
        else:
            cur.execute("""SELECT ticker, sector FROM companies
                           WHERE onboarding_status='discovered' AND NOT is_live
                           ORDER BY ticker""")
            rows = cur.fetchall()
            todo = [t for t, s in rows if s not in ("bank", "insurance")]
            skipped += [f"{t} (bankovni/osig. obrazac — v2)" for t, s in rows
                        if s in ("bank", "insurance")]
        for tick in todo:
            try:
                d = feed("financialReports", ticker=tick,
                         date_from="2025-01-01", date_to=today)
                rep = pick_report(d.get("items") or [])
                if not rep:
                    skipped.append(f"{tick} (nema 4Q/1Y XLSX na EHO)")
                    continue
                path = f"{SCRATCH}/{tick}.xlsx"
                r = requests.get(rep["documentLink"], timeout=120, verify=_verify())
                r.raise_for_status()
                open(path, "wb").write(r.content)
                parsed = parse_tfi(path)
                if not parsed or not parsed["items"].get("total_assets"):
                    skipped.append(f"{tick} (obrazac nije TFI-POD ili prazan)")
                    continue
                fy = rep["year"] if rep["period"] != "4Q" else rep["year"]
                solo_note = ("" if rep["consolidated"] else
                             " | SOLO obrazac (izdavatelj bez konsolidiranog TFI-ja)")
                extraction = {
                    "meta": {"company_ticker": tick, "fiscal_year": fy,
                             "period_type": "annual", "basis": "consolidated",
                             "audited": False, "cumulative": True,
                             "currency": "EUR", "reporting_scale": 1},
                    "items": [
                        {"item": k, "value_raw": v, "confidence": 0.9 if k != "ebit" else 0.85,
                         "source_page": (f"TFI {rep['period']} {rep['year']} XLSX, "
                                         f"{parsed['src'][k]}{solo_note} — strojno parsirano "
                                         f"(AOP obrazac), nerevidirani kumulativ")}
                        for k, v in parsed["items"].items()
                    ],
                }
                fid = load_extraction(conn, extraction,
                                      source_url=rep["documentLink"],
                                      doc_type="financial_report",
                                      published_at=rep.get("publishDate"))
                conn.commit()
                res = validate_filing(conn, fid)
                cur.execute("SELECT company_id FROM filings WHERE id=%s", (fid,))
                cid = cur.fetchone()[0]
                cur.execute("SELECT sector FROM companies WHERE id=%s", (cid,))
                sector = cur.fetchone()[0]
                blocking, _notes = core_gate(conn, fid, sector or "")
                if res["status"] == "validated" and not blocking:
                    cur.execute("""UPDATE companies SET is_live=TRUE,
                                   onboarding_status='live' WHERE id=%s""", (cid,))
                    conn.commit()
                    promoted.append(tick)
                    print(f"[tfi] {tick}: filing {fid} VALIDATED, gate čist -> LIVE "
                          f"({len(parsed['items'])} stavki)")
                else:
                    fails = [x["detail"][:80] for x in res["results"]
                             if x["status"] == "fail"] + blocking
                    review.append(f"{tick}: {'; '.join(fails[:2])}")
                    print(f"[tfi] {tick}: filing {fid} NEEDS_REVIEW — {fails[:2]}")
            except Exception as e:  # noqa: BLE001 — jedna firma ne ruši batch
                conn.rollback()
                skipped.append(f"{tick} (greška: {str(e)[:60]})")
                print(f"[tfi] {tick}: GREŠKA {e}")
    print(f"\nPROMOVIRANO ({len(promoted)}):", " ".join(promoted))
    print(f"NEEDS_REVIEW ({len(review)}):")
    for x in review:
        print("  -", x)
    print(f"PRESKOČENO ({len(skipped)}):")
    for x in skipped:
        print("  -", x)
    return 0


if __name__ == "__main__":
    sys.exit(main())
