"""M-FOND: mirovinski fondovi (OMF) — vrijednosti obračunskih jedinica i
Mirex, izvor HANFA (javne objave, MJESEČNI ritam — zaseban workflow, ne
dio dnevnog EOD-a).

VAŽNO o okruženju: hanfa.hr NIJE na mrežnoj allowlisti Claude Code
okruženja, pa se dohvat ovdje ne može testirati — fetch_hanfa() je pisan
za GitHub Actions runner (otvorena mreža). Parser je STROG: na nepoznat
format baca grešku s uputom (workflow tada otvara issue) — radije pad s
razlogom nego kriva brojka ("ništa izmišljeno"). CLI podržava i ručni
uvoz datoteke (--import-file) za prvi run / popravak formata.

Matching OMF-ova u shareholders tablici (dokumentirana pravila, koristi ih
build_fondovi.py):
  1. odbaci custodian prefiks (tekst prije zadnjeg '/')
  2. normaliziraj (velika slova, ligature ﬀ->FF, dijakritici)
  3. obitelj: 'AZ' | 'ERSTE PLAVI' | 'PBZ CO'/'PBZ CROATIA OSIGURANJE' |
     'RAIFFEISEN'; mora sadržavati 'OMF' ili 'OBVEZNI MIROVINSKI'
     (dobrovoljni fondovi se NE broje u OMF prikaz)
  4. kategorija: 'KATEGORIJ(A|E) A/B/C'
"""
from __future__ import annotations

import re
import unicodedata

FUNDS = ["AZ", "Erste Plavi", "PBZ CO", "Raiffeisen"]
CATEGORIES = ["A", "B", "C"]


def ensure_tables(conn) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fund_units (
                fund TEXT NOT NULL,        -- AZ | Erste Plavi | PBZ CO | Raiffeisen
                category TEXT NOT NULL,    -- A | B | C
                value_date DATE NOT NULL,
                unit_value NUMERIC NOT NULL,
                source TEXT,
                PRIMARY KEY (fund, category, value_date));
            CREATE TABLE IF NOT EXISTS mirex (
                category TEXT NOT NULL,    -- A | B | C (Mirex indeksi po kategoriji)
                value_date DATE NOT NULL,
                value NUMERIC NOT NULL,
                source TEXT,
                PRIMARY KEY (category, value_date));
            -- M-FOND3: neto imovina (AUM) po fondu/kategoriji + broj članova,
            -- HANFA mjesečna statistika OMF-ova. Bez ovoga se ne može ocijeniti
            -- veličina fonda (mali fond s 1% neke firme vs veliki fond).
            CREATE TABLE IF NOT EXISTS fund_aum (
                fund TEXT NOT NULL,        -- AZ | Erste Plavi | PBZ CO | Raiffeisen
                category TEXT NOT NULL,    -- A | B | C
                value_date DATE NOT NULL,
                net_assets_eur NUMERIC,    -- neto imovina fonda (EUR)
                members INTEGER,           -- broj članova (ako HANFA objavi)
                source TEXT,
                PRIMARY KEY (fund, category, value_date));
        """)
    conn.commit()


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.upper().strip()


def match_omf(holder_name: str) -> tuple[str, str] | None:
    """(obitelj, kategorija) iz naziva dioničara ili None (dokumentirana
    pravila u docstringu modula). Dobrovoljni fondovi -> None."""
    n = _norm(holder_name.split("/")[-1])
    if "OMF" not in n and "OBVEZNI MIROVINSKI" not in n:
        return None
    fam = None
    if re.search(r"\bAZ\b", n):
        fam = "AZ"
    elif "ERSTE PLAVI" in n:
        fam = "Erste Plavi"
    elif "PBZ CO" in n or "PBZ CROATIA OSIGURANJE" in n:
        fam = "PBZ CO"
    elif "RAIFFEISEN" in n:
        fam = "Raiffeisen"
    if not fam:
        return None
    m = re.search(r"KATEGORIJ[AE]\s*([ABC])\b", n)
    return (fam, m.group(1) if m else "B")  # bez oznake: B je najveća/default


def import_rows(conn, rows: list[dict], source: str) -> int:
    """Idempotentan BATCH upis jedinica i MIREX-a (ON CONFLICT DO NOTHING).
    c-03 nosi punu DNEVNU povijest od 2002. (~90k redova) — red-po-red
    upis preko mreže bi trajao desetke minuta (ista zamka kao db-sync #1),
    pa ide execute_values u serijama."""
    from psycopg2.extras import execute_values
    ensure_tables(conn)
    units = [r for r in rows if r.get("kind", "unit") == "unit"]
    mirex = [r for r in rows if r.get("kind") == "mirex"]
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM fund_units")
        before_u = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mirex")
        before_m = cur.fetchone()[0]
        if units:
            execute_values(cur,
                """INSERT INTO fund_units (fund, category, value_date, unit_value, source)
                   VALUES %s ON CONFLICT (fund, category, value_date) DO NOTHING""",
                [(r["fund"], r["category"], r["value_date"], r["unit_value"], source)
                 for r in units], page_size=2000)
        if mirex:
            # rekonstrukcija umjesto dopune: prvi uvoz (17.07.2026.) je
            # MIREX prije 2023. krivo preračunao u EUR (indeks se ne
            # preračunava) — ON CONFLICT DO NOTHING te redove nikad ne bi
            # ispravio, a datoteka ionako uvijek nosi PUNU povijest
            cur.execute("DELETE FROM mirex WHERE source LIKE 'HANFA statistika%%'")
            execute_values(cur,
                """INSERT INTO mirex (category, value_date, value, source)
                   VALUES %s ON CONFLICT (category, value_date) DO NOTHING""",
                [(r["category"], r["value_date"], r["value"], source)
                 for r in mirex], page_size=2000)
        cur.execute("SELECT COUNT(*) FROM fund_units")
        after_u = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mirex")
        after_m = cur.fetchone()[0]
    conn.commit()
    return (after_u - before_u) + (after_m - before_m)


def fetch_hanfa(conn, log=print) -> int:
    """Dohvat zadnjih objavljenih vrijednosti jedinica s HANFA-e.
    Pokreće se iz MJESEČNOG workflowa (GitHub Actions — otvorena mreža).
    STROG parser redova (nikad kriva brojka), ali TOLERANTAN prema
    strukturi stranice: HANFA datoteke poslužuje kroz /getfile/?fileId=N
    (bez .xlsx u URL-u; prvi run 17.07.2026. pao na 404 stare putanje i
    krivom href obrascu). Kandidati se probaju redom dok jedan ne prođe
    parser; nijedan -> RuntimeError s razlogom.
    Vraća broj NOVIH zapisa (0 = nema novih objava, idempotentno)."""
    import requests
    listings = (
        "https://www.hanfa.hr/statistika/mirovinski-fondovi/",
        "https://www.hanfa.hr/statistika/",
    )
    html, listing = None, None
    for cand_listing in listings:
        try:
            r = requests.get(cand_listing, timeout=90,
                             headers={"User-Agent": "Mozilla/5.0 (podaci; burzovnilist.com)"})
            r.raise_for_status()
            html, listing = r.text, cand_listing
            break
        except Exception as e:  # noqa: BLE001 — probaj sljedeći listing
            log(f"[hanfa] {cand_listing}: {type(e).__name__}: {e}")
    if html is None:
        raise RuntimeError(
            "HANFA: nijedna listing stranica nije dostupna "
            f"({', '.join(listings)}) — provjeri strukturu sajta")
    # <a href="..."> čiji href ILI tekst upućuje na statistiku mirovinskih
    # fondova; datoteke su .xlsx ILI /getfile/?fileId=N
    cands = []
    for m in re.finditer(r'<a\b[^>]*href="([^"]+)"[^>]*>(.*?)</a>', html, re.I | re.S):
        href, text = m.group(1), re.sub(r"<[^>]+>", " ", m.group(2))
        if not re.search(r"\.xlsx|/getfile/", href, re.I):
            continue
        blob = f"{href} {text}"
        if re.search(r"mirovin|omf|mirex|obra[čc]unsk", blob, re.I):
            cands.append(href if href.startswith("http")
                         else f"https://www.hanfa.hr{href}")
    if not cands:
        raise RuntimeError(
            f"HANFA: nijedna XLSX/getfile poveznica s 'mirovin' na {listing} "
            "— provjeri strukturu stranice i prilagodi fetch_hanfa()")
    # jedinice i MIREX žive u c-03 (neto imovina i Mirex) — nju prvo
    cands.sort(key=lambda u: 0 if re.search(r"mirex|neto_imovina", u, re.I) else 1)
    errors = []
    for url in cands[:6]:
        try:
            raw = requests.get(url, timeout=120).content
            if not raw.startswith(b"PK"):     # XLSX = zip; sve drugo preskoči
                errors.append(f"{url}: nije XLSX (magic bytes)")
                continue
            rows = parse_hanfa_xlsx(raw)
            log(f"[hanfa] parsirano {len(rows)} redova iz {url}")
            n = import_rows(conn, rows, f"HANFA statistika ({url})")
            # M-FOND3: iz istog workbooka pokušaj i neto imovinu (AUM) —
            # NEobavezno: ako list neto imovine nije prepoznat, ne ruši uvoz
            try:
                aum = parse_hanfa_aum(raw)
                if aum:
                    import_aum(conn, aum, f"HANFA neto imovina ({url})")
                    log(f"[hanfa] neto imovina: {len(aum)} fond/kategorija")
                else:
                    log("[hanfa] neto imovina: list nije prepoznat (AUM ostaje n/p)")
            except Exception as e:  # noqa: BLE001 — AUM je best-effort
                conn.rollback()
                log(f"[hanfa] neto imovina preskočena: {type(e).__name__}: {e}")
            return n
        except Exception as e:  # noqa: BLE001 — probaj sljedeći kandidat
            errors.append(f"{url}: {type(e).__name__}: {str(e)[:120]}")
    raise RuntimeError(
        "HANFA: nijedan od kandidata nije prošao strogi parser:\n  "
        + "\n  ".join(errors))


HRK_EUR = 7.5345   # fiksni konverzijski tečaj (uveden 1.1.2023.)


def _header_kind(cell: str):
    """('unit', obitelj, kat) | ('mirex', kat) | None iz teksta zaglavlja.
    Stvarni format (c-03, sheet 'C - OMF VOJ i MIREX', forenzika 17.07.2026.):
    'AZ\\nOMF A', 'Erste Plavi\\nOMF A', 'PBZ CO OMF A', 'Raiffeisen OMF A',
    'MIREX  A' — višak whitespacea/newlinea se normalizira."""
    n = _norm(re.sub(r"\s+", " ", str(cell))).strip()
    m = re.match(r"^MIREX\s*([ABC])$", n)
    if m:
        return ("mirex", m.group(1))
    m = re.match(r"^(AZ|ERSTE PLAVI|PBZ CO|PBZ CROATIA OSIGURANJE|RAIFFEISEN)"
                 r"\s+OMF\s*([ABC])$", n)
    if m:
        fam = {"PBZ CROATIA OSIGURANJE": "PBZ CO"}.get(m.group(1), m.group(1))
        fam = {"AZ": "AZ", "ERSTE PLAVI": "Erste Plavi", "PBZ CO": "PBZ CO",
               "RAIFFEISEN": "Raiffeisen"}[fam]
        return ("unit", fam, m.group(2))
    return None


def parse_hanfa_xlsx(raw: bytes) -> list[dict]:
    """Parsiranje HANFA c-03 XLSX-a (sheet s vrijednostima obračunskih
    jedinica OMF-ova i MIREX-om). STROGO prema izmjerenoj strukturi:
    zaglavlje = red s ćelijom 'Datum' + >=3 stupca fond/MIREX; podaci =
    redovi s datumom u datumskom stupcu. Vrijednosti s datumom do
    31.12.2022. su u HRK (footnote u datoteci) -> preračun u EUR fiksnim
    tečajem. Ništa prepoznato -> RuntimeError (nikad kriva brojka)."""
    import datetime as dt
    import io

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out = []
    cutoff = dt.date(2022, 12, 31)
    for ws in wb.worksheets:
        # M-FOND3: preskoči list neto imovine (isti workbook, iste kolone
        # fondova) — inače bi iznosi od stotina milijuna završili kao
        # obračunske jedinice. Neto imovinu čita parse_hanfa_aum().
        if re.search(r"neto\s*imovin|(?<![a-z])imovin", _norm(ws.title), re.I):
            continue
        colmap, date_col = None, None
        for row in ws.iter_rows(values_only=True):
            if colmap is None:
                cand = {}
                dcol = None
                for i, c in enumerate(row):
                    if c is None:
                        continue
                    if "DATUM" in _norm(str(c)):
                        dcol = i
                    k = _header_kind(c)
                    if k:
                        cand[i] = k
                if dcol is not None and len(cand) >= 3:
                    colmap, date_col = cand, dcol
                continue
            d = row[date_col] if date_col < len(row) else None
            if not isinstance(d, dt.datetime | dt.date):
                continue
            d = d.date() if isinstance(d, dt.datetime) else d
            hrk = d <= cutoff
            for i, kind in colmap.items():
                v = row[i] if i < len(row) else None
                if not isinstance(v, int | float):
                    continue
                if kind[0] == "unit":
                    # jedinice su do 31.12.2022. u HRK (footnote u datoteci),
                    # od 2023. u EUR — preračun drži seriju kontinuiranom
                    out.append({"kind": "unit", "fund": kind[1],
                                "category": kind[2], "value_date": d,
                                "unit_value": float(v) / HRK_EUR if hrk else float(v)})
                else:
                    # MIREX je INDEKS kontinuiteta (baza 100 na početku rada
                    # kategorije) — HANFA ga NE re-denominira u euro, pa se
                    # NE preračunava (2026: MIREX B ~377 uz jedinice ~48 EUR)
                    out.append({"kind": "mirex", "category": kind[1],
                                "value_date": d, "value": float(v)})
    if not out:
        raise RuntimeError(
            "HANFA XLSX: nijedan red nije prepoznat kao OMF jedinica/MIREX — "
            "format se promijenio ili je datoteka kriva; prilagodi "
            "parse_hanfa_xlsx() (debug: scripts/debug_hanfa.py)")
    return out


# M-FOND3: neto imovina (AUM) po fondu/kategoriji. Neto imovina fonda mjeri
# se u desecima/stotinama milijuna EUR — nikad ne može biti obračunska
# jedinica (~10–50 EUR), pa magnitudni gate (>= AUM_MIN) sprječava zamjenu.
AUM_MIN = 1_000_000.0


AUM_ALL_FUNDS = "SVI"   # HANFA c-03 objavljuje neto imovinu PO KATEGORIJI
#                         (svi OMF-ovi zajedno), ne po pojedinom fondu


def parse_hanfa_aum(raw: bytes) -> list[dict]:
    """NAJNOVIJA neto imovina PO KATEGORIJI iz HANFA c-03 workbooka.
    IZMJERENA struktura (debug-hanfa 19.07.2026., sheet 'C-3 imovina
    MF(EUR)'): iznosi su U TISUĆAMA EUR; DATUMI SU U STUPCIMA (zaglavlje
    = red s >=3 datetime ćelija); REDOVI su kategorije — 'Obvezni
    mirovinski fondovi kategorije A/B/C' sa šifrom 1/2/3 u prvom stupcu.
    HANFA NE objavljuje neto imovinu po pojedinom fondu, pa se sprema
    fund=AUM_ALL_FUNDS po kategoriji. HRK listovi i dobrovoljni fondovi
    se preskaču. STROGO: bez jasnog signala vraća [] — nikad kriva
    brojka."""
    import datetime as dt
    import io

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    latest: dict[str, tuple[dt.date, float]] = {}
    sifra_cat = {1: "A", 2: "B", 3: "C"}
    for ws in wb.worksheets:
        tn = _norm(ws.title)
        if "IMOVIN" not in tn or "HRK" in tn:
            continue                      # samo EUR list(ovi) neto imovine
        date_cols = None
        for row in ws.iter_rows(values_only=True):
            if date_cols is None:
                cand = {i: (c.date() if isinstance(c, dt.datetime) else c)
                        for i, c in enumerate(row)
                        if isinstance(c, dt.datetime | dt.date)}
                if len(cand) >= 3:
                    date_cols = cand
                continue
            label_cell = next((c for c in row[:3] if isinstance(c, str)), None)
            if label_cell is None:
                continue
            ln = _norm(label_cell)
            if "OBVEZNI MIROVINSKI" not in ln or ln.startswith("UKUPNO"):
                continue                  # dobrovoljni/ukupno se ne broje
            m = re.search(r"KATEGORIJ\w*\s+([ABC])\b", ln)
            cat = m.group(1) if m else sifra_cat.get(
                row[0] if isinstance(row[0], int) else None)
            if cat is None:
                continue
            for i, d in date_cols.items():
                v = row[i] if i < len(row) else None
                if not isinstance(v, int | float):
                    continue
                val = float(v) * 1000.0   # datoteka: 'u tisućama EUR'
                if val < AUM_MIN:
                    continue
                if cat not in latest or d > latest[cat][0]:
                    latest[cat] = (d, val)
    return [{"fund": AUM_ALL_FUNDS, "category": c, "value_date": d,
             "net_assets_eur": v} for c, (d, v) in latest.items()]


def parse_nav_lines(text: str) -> list[dict]:
    """M-FOND5 (opcija C): ručni/poluautomatski unos NETO IMOVINE POJEDINOG
    fonda. Format po retku: 'Fond;Kategorija;YYYY-MM-DD;iznos_eur'
    (npr. 'AZ;B;2026-06-30;9500000000'). Iznos je PUNI iznos u EUR (ne
    tisuće/milijuni); dopuštene su točke/razmaci kao separatori tisućica i
    zarez kao decimalni. STROGA validacija — kriva linija ruši cijeli unos
    s jasnom porukom (radije ništa nego kriva brojka)."""
    import datetime as dt
    rows = []
    for ln_no, raw_ln in enumerate(text.strip().splitlines(), 1):
        ln = raw_ln.strip()
        if not ln or ln.startswith("#"):
            continue
        parts = [p.strip() for p in ln.split(";")]
        if len(parts) != 4:
            raise ValueError(f"linija {ln_no}: očekujem 4 polja "
                             f"'Fond;Kategorija;Datum;Iznos', dobio: {ln!r}")
        fund, cat, d_s, v_s = parts
        if fund not in FUNDS:
            raise ValueError(f"linija {ln_no}: nepoznat fond {fund!r} "
                             f"(dopušteno: {', '.join(FUNDS)})")
        cat = cat.upper()
        if cat not in CATEGORIES:
            raise ValueError(f"linija {ln_no}: kategorija mora biti A/B/C, ne {cat!r}")
        try:
            d = dt.date.fromisoformat(d_s)
        except ValueError:
            raise ValueError(f"linija {ln_no}: datum mora biti YYYY-MM-DD, ne {d_s!r}")
        v_clean = v_s.replace(" ", "")
        if "," in v_clean:                 # HR zapis: točke tisućice, zarez dec.
            v_clean = v_clean.replace(".", "").replace(",", ".")
        elif v_clean.count(".") > 1:       # više točaka = separatori tisućica
            v_clean = v_clean.replace(".", "")
        try:
            val = float(v_clean)
        except ValueError:
            raise ValueError(f"linija {ln_no}: iznos nije broj: {v_s!r}")
        if not (AUM_MIN <= val <= 50e9):
            raise ValueError(
                f"linija {ln_no}: iznos {val:,.0f} EUR izvan razumnog raspona "
                f"[{AUM_MIN:,.0f}, 50 mlrd] — iznos mora biti PUNI EUR iznos")
        rows.append({"fund": fund, "category": cat, "value_date": d,
                     "net_assets_eur": val})
    if not rows:
        raise ValueError("nijedna valjana linija u unosu")
    return rows


def import_aum(conn, rows: list[dict], source: str) -> int:
    """Idempotentan upsert neto imovine (najnoviji snapshot po fondu/kat.)."""
    if not rows:
        return 0
    from psycopg2.extras import execute_values
    ensure_tables(conn)
    with conn.cursor() as cur:
        execute_values(cur,
            """INSERT INTO fund_aum
                 (fund, category, value_date, net_assets_eur, members, source)
               VALUES %s
               ON CONFLICT (fund, category, value_date)
               DO UPDATE SET net_assets_eur = EXCLUDED.net_assets_eur,
                             source = EXCLUDED.source""",
            [(r["fund"], r["category"], r["value_date"],
              r.get("net_assets_eur"), r.get("members"), source)
             for r in rows])
    conn.commit()
    return len(rows)


if __name__ == "__main__":
    import argparse
    import sys
    sys.path.insert(0, ".")
    from .db import get_conn
    ap = argparse.ArgumentParser()
    ap.add_argument("--import-file", help="ručni uvoz HANFA XLSX datoteke")
    ap.add_argument("--import-nav", help="unos NAV-a po fondu: datoteka s "
                    "linijama 'Fond;Kategorija;YYYY-MM-DD;iznos_eur'")
    ap.add_argument("--import-nav-lines", help="isto kao --import-nav, ali "
                    "linije direktno kao argument (za CI workflow)")
    ap.add_argument("--nav-source", default="mjesečni izvještaj fonda (ručni unos)",
                    help="izvor uz NAV unos (piše se uz svaku brojku)")
    a = ap.parse_args()
    with get_conn() as conn:
        ensure_tables(conn)
        if a.import_nav or a.import_nav_lines:
            text = (open(a.import_nav, encoding="utf-8").read()
                    if a.import_nav else a.import_nav_lines)
            rows = parse_nav_lines(text)
            n = import_aum(conn, rows, a.nav_source)
            for r in rows:
                print(f"  {r['fund']} {r['category']} {r['value_date']}: "
                      f"{r['net_assets_eur'] / 1e6:,.1f} M EUR")
            print(f"NAV unos: {n} zapisa (izvor: {a.nav_source})")
        elif a.import_file:
            raw = open(a.import_file, "rb").read()
            rows = parse_hanfa_xlsx(raw)
            print(f"novo: {import_rows(conn, rows, f'HANFA XLSX (ručni uvoz: {a.import_file})')}")
            aum = parse_hanfa_aum(raw)
            print(f"neto imovina: {import_aum(conn, aum, f'HANFA XLSX AUM (ručni uvoz: {a.import_file})')}")
        else:
            print(f"novo: {fetch_hanfa(conn)}")
